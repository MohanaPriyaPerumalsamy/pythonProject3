import openpyxl

wbBook=openpyxl.load_workbook("C:\\Users\\ADMIN\\OneDrive\\Desktop\\pythonexceloperation.xlsx")
wbsheet=wbBook.active
#cell=wbsheet.cell(row=1,column=1)
#print(cell.value)
#cell2=wbsheet.cell(row=1,column=2)
#wbsheet.cell(row=1,column=2).value="mohana"
#print(cell2.value)
a=wbsheet.cell(row=1,column=2)
print(a.value)
b=wbsheet.cell(row=1,column=3)
print(b.value)
c=a.value+b.value
print(c)
wbsheet.cell(row=1,column=4).value=c
s=wbsheet.cell(row=2,column=2).value - wbsheet.cell(row=2,column=3).value
wbsheet.cell(row=2,column=4).value=s
m=wbsheet.cell(row=3,column=2).value * wbsheet.cell(row=3,column=3).value
wbsheet.cell(row=3,column=4).value=m
d=wbsheet.cell(row=4,column=2).value / wbsheet.cell(row=4,column=3).value
wbsheet.cell(row=4,column=4).value=d
mo=wbsheet.cell(row=5,column=2).value % wbsheet.cell(row=5,column=3).value
wbsheet.cell(row=5,column=4).value=mo
#print(c.value)
wbBook.save("C:\\Users\\ADMIN\\OneDrive\\Desktop\\pythonexceloperation.xlsx")

